"""
Serviço de Conciliação de Recebimentos — LLE Índices (Processo 1).

Entradas: extrato bancário + base de recebimentos (com "Identificador do pagamento").
Classifica os recebimentos/movimentos e apura o PIX que sobra (para o Monday).

Resumo (totais, no topo):
  QR Boleto, PDV, CobCloud, Boleto Cód Barras, Transferência entre contas,
  TED, DOC, Sobrando (Monday), Despesas, Tarifas, Aplicação.
Detalhe: apenas PIX sobrando (+ ambíguos). Despesas/Tarifas/Aplicação/
Transferência/TED/DOC entram só como soma no resumo.
"""
from __future__ import annotations
import io
import re
from collections import Counter, defaultdict
from typing import Optional

import pandas as pd

QR_BOLETO = "QR Boleto"
PDV = "PDV"
COBCLOUD = "CobCloud"
BOLETO_CB = "Boleto Cód Barras"
TRANSF = "Transferência entre contas"
TED = "TED"
DOC = "DOC"
SOBRA = "Sobrando (Monday)"
DESPESA = "Despesas"
TARIFA = "Tarifas"
APLICACAO = "Aplicação"

ID_COBCLOUD = "3455/000554357"
ID_BOLETO_CB = "3455/003475395"

# ordem de exibição no resumo
ORDEM_RESUMO = [QR_BOLETO, PDV, COBCLOUD, BOLETO_CB, TRANSF, TED, DOC,
                SOBRA, DESPESA, TARIFA, APLICACAO]


def _categoria_extrato(hist: str, valor: float) -> str:
    """Categoria de uma linha do extrato (exceto QR/PDV, que vêm da base)."""
    h = str(hist).upper()
    if ID_COBCLOUD in h:
        return COBCLOUD
    if ID_BOLETO_CB in h:
        return BOLETO_CB
    if "PIX RECEBIDO" in h and valor > 0:
        return "PIX"  # entra no casamento → sobrando
    if "TRANSFERENCIA ENTRE CONTAS" in h or "TRANSF ENTRE CONTAS" in h:
        return TRANSF
    if re.search(r"\bTED\b", h):
        return TED
    if re.search(r"\bDOC\b", h):
        return DOC
    if "TARIFA" in h:
        return TARIFA
    if "APLICA" in h:
        return APLICACAO
    if valor < 0:
        return DESPESA
    return "OUTROS"


def processar_conciliacao(arquivos: list[tuple[bytes, str]]) -> dict:
    res = {"resumo": [], "sobra": [], "sobra_ambigua": [], "total_sobra": 0.0,
           "erros": [], "arquivos": [], "ok": False}

    base_df = extrato_df = None
    faturados = []
    for bytes_arq, nome in arquivos:
        tipo = _detectar(bytes_arq, nome)
        if tipo == "BASE":
            base_df = _ler_base(bytes_arq, nome); res["arquivos"].append(f"{nome} → Base recebimentos")
        elif tipo == "EXTRATO":
            extrato_df = _ler_extrato(bytes_arq, nome); res["arquivos"].append(f"{nome} → Extrato bancário")
        elif tipo == "FATURADOS":
            faturados = ler_faturados(bytes_arq, nome); res["arquivos"].append(f"{nome} → Faturados (Monday)")
        else:
            res["erros"].append(f"'{nome}': não reconhecido (extrato, base com 'Identificador do pagamento', ou faturados com 'Vlr do Desdobramento').")

    if base_df is None:
        res["erros"].append("Faltou a **base de recebimentos** (com 'Identificador do pagamento').")
    if extrato_df is None:
        res["erros"].append("Faltou o **extrato bancário**.")
    if base_df is None or extrato_df is None:
        return res

    # --- base: QR Boleto e PDV ---
    base_df["cat"] = base_df["identificador"].apply(_classificar_id)
    base_df["valor"] = pd.to_numeric(base_df["valor"], errors="coerce").round(2)
    qtd_qr = int((base_df["cat"] == QR_BOLETO).sum())
    val_qr = round(base_df[base_df["cat"] == QR_BOLETO]["valor"].sum(), 2)
    qtd_pdv = int((base_df["cat"] == PDV).sum())
    val_pdv = round(base_df[base_df["cat"] == PDV]["valor"].sum(), 2)

    # --- extrato: categorias ---
    extrato_df["valor"] = pd.to_numeric(extrato_df["valor"], errors="coerce")
    extrato_df["cat"] = [_categoria_extrato(h, v) for h, v in zip(extrato_df["historico"], extrato_df["valor"])]
    tot = defaultdict(lambda: [0, 0.0])  # cat -> [qtd, soma]
    for _, r in extrato_df.iterrows():
        if r["cat"] in ("PIX", "OUTROS"):
            continue
        tot[r["cat"]][0] += 1
        tot[r["cat"]][1] += float(r["valor"] or 0)

    # --- PIX positivo → casa com automáticos (QR+PDV) → sobrando ---
    pix = extrato_df[(extrato_df["cat"] == "PIX")].copy()
    pix["valor"] = pd.to_numeric(pix["valor"], errors="coerce").round(2)
    auto = [round(v, 2) for v in base_df[base_df["cat"].isin([QR_BOLETO, PDV])]["valor"].dropna().tolist()]
    auto_cont = Counter(auto)
    pix_por_valor = defaultdict(list)
    for _, r in pix.iterrows():
        v = round(float(r["valor"]), 2)
        pix_por_valor[v].append({"data": str(r["data"]), "historico": str(r["historico"]).strip(), "valor": v})
    sobra, sobra_ambigua = [], []
    for v, rows in pix_por_valor.items():
        a_v = auto_cont.get(v, 0)
        leftover = max(0, len(rows) - a_v)
        if leftover == 0:
            continue
        if a_v == 0:
            sobra.extend(rows)
        else:
            sobra_ambigua.append({"valor": v, "conta": leftover, "candidatos": rows})
    total_sobra = round(sum(s["valor"] for s in sobra) + sum(g["valor"] * g["conta"] for g in sobra_ambigua), 2)
    qtd_sobra = len(sobra) + sum(g["conta"] for g in sobra_ambigua)

    # identifica a origem (cód parceiro / parceiro / nota) do que sobrou — continua sendo sobra
    if faturados:
        _identificar_sobra(sobra, sobra_ambigua, faturados)

    def cat_tot(nome):
        return {"tipo": nome, "qtd": int(tot[nome][0]), "valor": round(tot[nome][1], 2)}

    res["resumo"] = [
        {"tipo": QR_BOLETO, "qtd": qtd_qr, "valor": val_qr},
        {"tipo": PDV, "qtd": qtd_pdv, "valor": val_pdv},
        cat_tot(COBCLOUD), cat_tot(BOLETO_CB), cat_tot(TRANSF), cat_tot(TED), cat_tot(DOC),
        {"tipo": SOBRA, "qtd": qtd_sobra, "valor": total_sobra},
        cat_tot(DESPESA), cat_tot(TARIFA), cat_tot(APLICACAO),
    ]
    res["sobra"] = sorted(sobra, key=lambda s: -s["valor"])
    res["sobra_ambigua"] = sorted(sobra_ambigua, key=lambda g: -g["valor"])
    res["total_sobra"] = total_sobra
    res["ok"] = True
    return res


# ============================================================
# DETECÇÃO E LEITURA
# ============================================================

def _detectar(bytes_arq: bytes, nome: str) -> str:
    try:
        amostra = pd.read_excel(io.BytesIO(bytes_arq), sheet_name=0, header=None, nrows=6)
    except Exception:
        return "DESCONHECIDO"
    textos = " ".join(str(v).upper() for _, row in amostra.iterrows() for v in row.values if pd.notna(v))
    if "IDENTIFICADOR DO PAGAMENTO" in textos:
        return "BASE"
    if "VLR DO DESDOBRAMENTO" in textos or ("DESDOBRAMENTO" in textos and "PARCEIRO" in textos):
        return "FATURADOS"
    if "SALDO" in textos and ("HISTÓRICO" in textos or "HISTORICO" in textos):
        return "EXTRATO"
    if "AGENCIA" in textos and "CONTA" in textos:
        return "EXTRATO"
    return "DESCONHECIDO"


def _ler_base(bytes_arq: bytes, nome: str) -> pd.DataFrame:
    df = pd.read_excel(io.BytesIO(bytes_arq), sheet_name=0)
    df.columns = [str(c).strip() for c in df.columns]
    idc = _col(df, ["identificador do pagamento", "identificador"])
    valc = _col(df, ["valor"])
    out = pd.DataFrame({"identificador": df[idc], "valor": df[valc]})
    return out[out["valor"].notna()].copy()


def _ler_extrato(bytes_arq: bytes, nome: str) -> pd.DataFrame:
    raw = pd.read_excel(io.BytesIO(bytes_arq), sheet_name=0, header=None)
    hdr = 0
    for i in range(min(6, len(raw))):
        linha = [str(v).strip().lower() for v in raw.iloc[i].tolist()]
        if any("histórico" in c or "historico" in c for c in linha) and any("valor" in c for c in linha):
            hdr = i; break
    df = pd.read_excel(io.BytesIO(bytes_arq), sheet_name=0, header=hdr)
    df.columns = [str(c).strip() for c in df.columns]
    col_data = _col(df, ["data"])
    col_hist = _col(df, ["histórico", "historico"])
    col_val = _col(df, ["valor (r$)", "valor"])
    out = pd.DataFrame({"data": df[col_data], "historico": df[col_hist], "valor": df[col_val]})
    return out[out["data"].astype(str).str.contains(r"\d{2}/\d{2}", na=False)].copy()


def _classificar_id(x) -> str:
    s = str(x).strip().upper()
    if s.startswith("YKP"):
        return QR_BOLETO
    if s.startswith(ID_COBCLOUD):
        return COBCLOUD
    if s.startswith(ID_BOLETO_CB) or s.startswith("3455/"):
        return BOLETO_CB
    if s.startswith("010"):
        return PDV
    return "SEM ID"


def _col(df, nomes) -> Optional[str]:
    low = [str(c).strip().lower() for c in df.columns]
    for n in nomes:
        for i, c in enumerate(low):
            if c == n:
                return df.columns[i]
    for n in nomes:
        for i, c in enumerate(low):
            if n in c:
                return df.columns[i]
    return None


# ============================================================
# PROCESSO 2 — cruza o que sobrou (Monday) com os FATURADOS
# ============================================================

def ler_faturados(bytes_arq: bytes, nome: str) -> list[dict]:
    """Lê a planilha de faturados (Monday): código parceiro, parceiro, nota, valor do desdobramento."""
    raw = pd.read_excel(io.BytesIO(bytes_arq), sheet_name=0, header=None, engine=_engine(nome))
    hdr = 0
    for i in range(min(8, len(raw))):
        linha = [str(v).strip().lower() for v in raw.iloc[i].tolist()]
        if any("parceiro" in c for c in linha) and any("desdobramento" in c for c in linha):
            hdr = i; break
    df = pd.read_excel(io.BytesIO(bytes_arq), sheet_name=0, header=hdr, engine=_engine(nome))
    df.columns = [str(c).strip() for c in df.columns]
    c_cod = _col(df, ["parceiro"])
    c_nota = _col(df, ["nro nota", "nota"])
    c_nome = _col(df, ["nome parceiro", "nome"])
    c_val = _col(df, ["vlr do desdobramento", "desdobramento"])
    df = df[pd.to_numeric(df[c_cod], errors="coerce").notna()].copy()
    out = []
    for _, r in df.iterrows():
        out.append({
            "cod": int(float(r[c_cod])),
            "nota": str(r[c_nota]).replace(".0", "") if pd.notna(r[c_nota]) else "",
            "nome": str(r[c_nome]) if c_nome else "",
            "valor": round(float(pd.to_numeric(r[c_val], errors="coerce") or 0), 2),
        })
    return out


def _identificar_sobra(sobra: list[dict], sobra_ambigua: list[dict], faturados: list[dict]):
    """Preenche cód parceiro / parceiro / nota nas sobras que casam por valor (um-para-um).
    Continua sendo sobra — só identifica a origem."""
    fat = list(faturados)
    usado = [False] * len(fat)

    def achar(v):
        for i, fr in enumerate(fat):
            if not usado[i] and round(float(fr["valor"]), 2) == round(v, 2):
                usado[i] = True
                return fr
        return None

    for s in sorted(sobra, key=lambda s: -s["valor"]):
        fr = achar(s["valor"])
        if fr:
            s["cod"] = fr["cod"]; s["nome"] = fr["nome"]; s["nota"] = fr["nota"]
    for g in sorted(sobra_ambigua, key=lambda g: -g["valor"]):
        for _ in range(g["conta"]):
            fr = achar(g["valor"])
            if fr:
                g.setdefault("ident", []).append(fr)


def _engine(nome: str):
    return "xlrd" if str(nome).lower().endswith(".xls") else None


# ============================================================
# PLANILHA (resumo no topo + só o detalhe do que sobra)
# ============================================================

def gerar_xlsx_conciliacao(resumo: list[dict], sobra: list[dict], data_label: str,
                           sobra_ambigua: list[dict] | None = None, **_ignore) -> bytes:
    sobra_ambigua = sobra_ambigua or []
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    AZUL = "041747"; VERM = "DC3545"; CINZA = "6C757D"
    COR = {QR_BOLETO: "0F8C3B", PDV: "0071FE", COBCLOUD: "041747", BOLETO_CB: "0071FE",
           TRANSF: "6C757D", TED: "6C757D", DOC: "6C757D", SOBRA: "DC3545",
           DESPESA: "B00020", TARIFA: "B00020", APLICACAO: "7B5800"}
    thin = Side(style="thin", color="D9DCE3")

    def F(**k): return Font(name="Arial", **k)

    wb = Workbook(); ws = wb.active; ws.title = "Conciliação"
    ws["A1"] = f"Conciliação de Recebimentos — {data_label}"; ws["A1"].font = F(size=15, bold=True, color=AZUL)
    ws["A2"] = "Resumo dos recebimentos/movimentos e detalhe do PIX sobrando (Monday)"; ws["A2"].font = F(size=9, color=CINZA)

    ws["A4"] = "RESUMO"; ws["A4"].font = F(size=11, bold=True, color=AZUL)
    for j, c in enumerate(["Tipo", "Qtd", "Valor (R$)"], 1):
        cell = ws.cell(row=5, column=j, value=c); cell.font = F(size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=AZUL); cell.alignment = Alignment(horizontal="left" if j == 1 else "center")
    for i, x in enumerate(resumo):
        rr = 6 + i
        ws.cell(row=rr, column=1, value=x["tipo"]).font = F(size=10, bold=True, color=COR.get(x["tipo"], AZUL))
        qc = ws.cell(row=rr, column=2, value=x["qtd"]); qc.font = F(size=10); qc.alignment = Alignment(horizontal="center")
        vc = ws.cell(row=rr, column=3, value=x["valor"]); vc.font = F(size=10); vc.number_format = '#,##0.00;[Red]-#,##0.00'
        for j in range(1, 4): ws.cell(row=rr, column=j).border = Border(bottom=thin)

    base = 6 + len(resumo) + 2
    ws.cell(row=base, column=1, value="DETALHE — PIX SOBRANDO (Monday)").font = F(size=11, bold=True, color=VERM)
    cab = ["#", "Data", "Histórico", "Cód Parceiro", "Parceiro", "Nota", "Valor (R$)"]
    for j, c in enumerate(cab, 1):
        cell = ws.cell(row=base + 1, column=j, value=c); cell.font = F(size=10, bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=VERM)
        cell.alignment = Alignment(horizontal="left" if c in ("Histórico", "Parceiro") else "center")
    for i, s in enumerate(sobra):
        rr = base + 2 + i
        ws.cell(row=rr, column=1, value=i + 1).font = F(size=10)
        ws.cell(row=rr, column=2, value=s["data"]).font = F(size=10)
        ws.cell(row=rr, column=3, value=s["historico"]).font = F(size=10)
        ws.cell(row=rr, column=4, value=(s.get("cod") or "")).font = F(size=10)
        ws.cell(row=rr, column=5, value=(s.get("nome") or "")).font = F(size=10)
        ws.cell(row=rr, column=6, value=(str(s.get("nota")) if s.get("nota") else "")).font = F(size=10)
        vc = ws.cell(row=rr, column=7, value=float(s["valor"])); vc.font = F(size=10); vc.number_format = '#,##0.00'
        for j in range(1, 8): ws.cell(row=rr, column=j).border = Border(bottom=thin)
    tr2 = base + 2 + len(sobra)
    ws.cell(row=tr2, column=6, value="Subtotal (claros)").font = F(size=10, bold=True, color=VERM)
    ws.cell(row=tr2, column=6).alignment = Alignment(horizontal="right")
    ts = ws.cell(row=tr2, column=7, value=f"=SUM(G{base+2}:G{tr2-1})" if sobra else 0)
    ts.font = F(size=10, bold=True, color=VERM); ts.number_format = '#,##0.00'

    linha = tr2
    total_ambiguo = round(sum(g["valor"] * g["conta"] for g in sobra_ambigua), 2)
    if sobra_ambigua:
        ab = tr2 + 2
        ws.cell(row=ab, column=1, value="PIX AMBÍGUOS — mesmo valor de QR/PIX (não somar; só um sobra)").font = F(size=11, bold=True, color="7B5800")
        for j, c in enumerate(["Valor (R$)", "Sobrando", "Candidatos (lado a lado)"], 1):
            cell = ws.cell(row=ab + 1, column=j, value=c); cell.font = F(size=10, bold=True, color="3A2D00")
            cell.fill = PatternFill("solid", fgColor="FAC318")
        for i, g in enumerate(sobra_ambigua):
            rr = ab + 2 + i
            vc = ws.cell(row=rr, column=1, value=float(g["valor"])); vc.font = F(size=10, bold=True); vc.number_format = '#,##0.00'
            ws.cell(row=rr, column=2, value=f'{g["conta"]} de {len(g["candidatos"])}').font = F(size=10, color="7B5800")
            for k, cand in enumerate(g["candidatos"]):
                ws.cell(row=rr, column=3 + k, value=f'{cand["data"]} · {cand["historico"]}').font = F(size=9)
        linha = ab + 2 + len(sobra_ambigua)

    gtot = linha + 1
    ws.cell(row=gtot, column=6, value="TOTAL sobrando (Monday)").font = F(size=11, bold=True, color=VERM)
    ws.cell(row=gtot, column=6).alignment = Alignment(horizontal="right")
    gt = ws.cell(row=gtot, column=7, value=round(sum(s["valor"] for s in sobra) + total_ambiguo, 2))
    gt.font = F(size=11, bold=True, color=VERM); gt.number_format = '#,##0.00'; gt.fill = PatternFill("solid", fgColor="FDE7E9")

    for j, w in zip(range(1, 8), [6, 12, 34, 13, 30, 12, 16]): ws.column_dimensions[get_column_letter(j)].width = w
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()
