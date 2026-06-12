"""
Tela Financeiro · Comparativos e Dashboard

Mostra:
  - Evolução mês a mês (12 meses do ano)
  - Comparativo ano vs ano (esse ano vs anterior)
  - Quebra por categoria (evolução ao longo do tempo)
  - Quebra por fornecedor (top 10)
  - Quebra por empresa LLE (PISA/KING/TRIO)
  - KPIs gerais
"""
from __future__ import annotations

import datetime

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st

from src.banco.conexao import obter_conexao
from src.utils.formatadores import formatar_brl, nome_mes
from src.utils.marca import AZUL_ESCURO, AMARELO, VERDE


def _carregar_todos_gastos() -> pd.DataFrame:
    """Busca todos os lançamentos como DataFrame."""
    sb = obter_conexao()
    res = sb.table("dados_financeiro_gasto").select("*").execute()
    if not res.data:
        return pd.DataFrame()
    df = pd.DataFrame(res.data)
    df["valor"] = df["valor"].astype(float)
    df["ano"] = df["mes_ano"].str[:4].astype(int)
    df["mes"] = df["mes_ano"].str[5:7].astype(int)
    return df


def renderizar_fin_comparativos(usuario):
    st.markdown(
        f"<h1 style='color:{AZUL_ESCURO}'>📈 Financeiro · Comparativos e Dashboard</h1>",
        unsafe_allow_html=True,
    )
    st.caption(
        "Veja a evolução mês a mês e compare anos pra orçar o próximo período."
    )

    df = _carregar_todos_gastos()
    if df.empty:
        st.info(
            "📭 Nenhum gasto cadastrado ainda. Vá em **💼 Upload de NFs** "
            "ou **✍️ Cadastro Manual** pra começar."
        )
        return

    anos_disponiveis = sorted(df["ano"].unique(), reverse=True)
    ano_atual = datetime.date.today().year

    # ─── Seletor de ano ─────────────────────────
    col_ano, col_comp = st.columns([1, 1])
    with col_ano:
        ano_foco = st.selectbox(
            "🗓️ Ano em análise",
            anos_disponiveis,
            index=anos_disponiveis.index(ano_atual) if ano_atual in anos_disponiveis else 0,
            key="fin_comp_ano",
        )
    with col_comp:
        anos_para_comp = [a for a in anos_disponiveis if a != ano_foco]
        if anos_para_comp:
            ano_compara = st.selectbox(
                "📊 Comparar com",
                anos_para_comp,
                index=0,
                key="fin_comp_outro",
            )
        else:
            ano_compara = None
            st.caption("Sem outro ano disponível pra comparação ainda.")

    df_ano = df[df["ano"] == ano_foco]
    df_compara = df[df["ano"] == ano_compara] if ano_compara else pd.DataFrame()

    # ─── KPIs principais ─────────────────────────
    st.markdown("### 📊 Visão geral")

    total_ano = df_ano["valor"].sum()
    total_compara = df_compara["valor"].sum() if not df_compara.empty else 0
    qtd_lanc = len(df_ano)
    fornecedores_ativos = df_ano["fornecedor_id"].nunique()

    delta_pct = None
    if total_compara > 0:
        delta_pct = (total_ano - total_compara) / total_compara * 100

    c1, c2, c3, c4 = st.columns(4)
    c1.metric(
        f"💰 Total {ano_foco}",
        formatar_brl(total_ano),
        delta=f"{delta_pct:+.1f}% vs {ano_compara}" if delta_pct is not None else None,
        delta_color="inverse",  # gasto maior = vermelho
    )
    c2.metric(
        "📊 Média/mês",
        formatar_brl(total_ano / max(df_ano["mes"].nunique(), 1)),
    )
    c3.metric("📄 Lançamentos", qtd_lanc)
    c4.metric("🏢 Fornecedores", fornecedores_ativos)

    st.markdown("---")

    # ─── Evolução mês a mês (ano em foco) ─────────────────────────
    st.markdown(f"### 📈 Evolução mensal — {ano_foco}")

    # Monta DF com 12 meses (preenche zero onde não tem dado)
    meses_completos = pd.DataFrame({"mes": range(1, 13)})
    por_mes_foco = (df_ano.groupby("mes")["valor"].sum().reindex(range(1, 13), fill_value=0)).reset_index()
    por_mes_foco["ano"] = ano_foco

    if not df_compara.empty:
        por_mes_compara = (
            df_compara.groupby("mes")["valor"].sum().reindex(range(1, 13), fill_value=0)
        ).reset_index()
        por_mes_compara["ano"] = ano_compara
        plot_data = pd.concat([por_mes_foco, por_mes_compara], ignore_index=True)
    else:
        plot_data = por_mes_foco

    plot_data["mes_nome"] = plot_data["mes"].apply(
        lambda m: nome_mes(f"2026-{m:02d}")[:3]
    )

    fig = px.bar(
        plot_data,
        x="mes_nome",
        y="valor",
        color="ano",
        barmode="group",
        text_auto=".2s",
        labels={"valor": "Total (R$)", "mes_nome": "Mês", "ano": "Ano"},
        color_discrete_map={
            ano_foco: AZUL_ESCURO,
            ano_compara: AMARELO if ano_compara else AZUL_ESCURO,
        },
    )
    fig.update_layout(
        height=380,
        showlegend=bool(ano_compara),
        margin=dict(t=10, b=10),
    )
    fig.update_traces(textposition="outside")
    st.plotly_chart(fig, use_container_width=True)

    # Tabela de variação mensal
    if not df_compara.empty:
        st.markdown("#### Variação mês a mês")
        var_rows = []
        for m in range(1, 13):
            v_foco = por_mes_foco[por_mes_foco["mes"] == m]["valor"].iloc[0] if not por_mes_foco.empty else 0
            v_comp = por_mes_compara[por_mes_compara["mes"] == m]["valor"].iloc[0] if not por_mes_compara.empty else 0
            if v_foco == 0 and v_comp == 0:
                continue
            var = ((v_foco - v_comp) / v_comp * 100) if v_comp > 0 else None
            var_rows.append({
                "Mês": nome_mes(f"2026-{m:02d}"),
                f"{ano_compara}": formatar_brl(v_comp),
                f"{ano_foco}": formatar_brl(v_foco),
                "Diferença": formatar_brl(v_foco - v_comp),
                "Variação": f"{var:+.1f}%" if var is not None else "—",
            })
        if var_rows:
            st.dataframe(pd.DataFrame(var_rows), use_container_width=True, hide_index=True)

    st.markdown("---")

    # ─── Por Categoria ─────────────────────────
    st.markdown(f"### 🏷️ Por Categoria — {ano_foco}")

    col_cat1, col_cat2 = st.columns([2, 3])

    por_cat = df_ano.groupby("categoria", dropna=False)["valor"].sum().sort_values(ascending=False)

    with col_cat1:
        st.markdown("**Total por categoria**")
        rows = []
        for cat, val in por_cat.items():
            rows.append({
                "Categoria": cat or "—",
                "Total": formatar_brl(val),
                "%": f"{val / total_ano * 100:.1f}%" if total_ano else "0%",
            })
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    with col_cat2:
        # Pizza
        if not por_cat.empty:
            fig_pizza = px.pie(
                values=por_cat.values,
                names=[c or "—" for c in por_cat.index],
                hole=0.45,
            )
            fig_pizza.update_traces(textposition="inside", textinfo="percent+label")
            fig_pizza.update_layout(height=350, margin=dict(t=10, b=10), showlegend=False)
            st.plotly_chart(fig_pizza, use_container_width=True)

    # Evolução das TOP 5 categorias no ano
    top5_cats = por_cat.head(5).index.tolist()
    if top5_cats:
        st.markdown("#### Evolução mensal das principais categorias")
        ev_data = (df_ano[df_ano["categoria"].isin(top5_cats)]
                   .groupby(["mes", "categoria"])["valor"].sum().reset_index())
        ev_data["mes_nome"] = ev_data["mes"].apply(lambda m: nome_mes(f"2026-{m:02d}")[:3])

        fig_ev = px.line(
            ev_data,
            x="mes_nome",
            y="valor",
            color="categoria",
            markers=True,
            labels={"valor": "Total (R$)", "mes_nome": "Mês", "categoria": "Categoria"},
        )
        fig_ev.update_layout(height=350, margin=dict(t=10, b=10))
        st.plotly_chart(fig_ev, use_container_width=True)

    st.markdown("---")

    # ═════════════════════════════════════════════════════════════
    # 🎯 VISÃO EXECUTIVA POR ÁREA DE NEGÓCIO
    # ═════════════════════════════════════════════════════════════
    st.markdown(f"### 🎯 Visão Executiva — pra onde está indo o dinheiro?")
    st.caption(
        "Gastos agrupados pelas **áreas de negócio da LLE**. "
        "Pra reclassificar fornecedores, vá em **🏢 Fornecedores**."
    )

    from src.banco.repo_financeiro import AREAS_NEGOCIO

    # Trata área não classificada
    df_ano_area = df_ano.copy()
    df_ano_area["area_negocio"] = df_ano_area["area_negocio"].fillna("⚠️ NÃO CLASSIFICADO")

    # Total por área
    por_area = (df_ano_area.groupby("area_negocio")["valor"]
                .sum().sort_values(ascending=False))

    if por_area.empty:
        st.info("Sem dados pra classificar.")
    else:
        # Cards macro (4 colunas)
        st.markdown("##### 📊 Distribuição por Área")
        cols_cards = st.columns(min(len(por_area), 5))
        for idx, (area, val) in enumerate(por_area.items()):
            with cols_cards[idx % len(cols_cards)]:
                cor = AREAS_NEGOCIO.get(area, {}).get("cor", "#616161")
                emoji = AREAS_NEGOCIO.get(area, {}).get("emoji", "❓")
                pct = (val / total_ano * 100) if total_ano else 0
                st.markdown(
                    f"""<div style='background:{cor}; color:white; padding:14px;
                    border-radius:8px; margin-bottom:10px; text-align:center;'>
                    <div style='font-size:28px;'>{emoji}</div>
                    <div style='font-size:13px; font-weight:600; margin:4px 0;'>{area}</div>
                    <div style='font-size:18px; font-weight:700;'>{formatar_brl(val)}</div>
                    <div style='font-size:12px; opacity:0.9;'>{pct:.1f}% do total</div>
                    </div>""",
                    unsafe_allow_html=True,
                )

        # 2 gráficos lado a lado: pizza por área + barras horizontais
        col_p, col_b = st.columns([1, 1])

        with col_p:
            st.markdown("##### 🥧 Proporção por Área")
            cores_area = [AREAS_NEGOCIO.get(a, {}).get("cor", "#616161") for a in por_area.index]
            fig_pie = px.pie(
                names=por_area.index,
                values=por_area.values,
                color=por_area.index,
                color_discrete_map={a: AREAS_NEGOCIO.get(a, {}).get("cor", "#616161") for a in por_area.index},
                hole=0.4,
            )
            fig_pie.update_traces(textposition='inside', textinfo='percent+label')
            fig_pie.update_layout(height=350, margin=dict(t=10, b=10), showlegend=False)
            st.plotly_chart(fig_pie, use_container_width=True)

        with col_b:
            st.markdown("##### 📊 Total por Área")
            df_bar = por_area.reset_index()
            df_bar.columns = ["Área", "Valor"]
            fig_bar = px.bar(
                df_bar, x="Valor", y="Área", orientation="h",
                color="Área",
                color_discrete_map={a: AREAS_NEGOCIO.get(a, {}).get("cor", "#616161") for a in por_area.index},
                text_auto=".2s",
                labels={"Valor": "R$", "Área": ""},
            )
            fig_bar.update_layout(
                height=350, margin=dict(t=10, b=10), showlegend=False,
                yaxis={"categoryorder": "total ascending"},
            )
            st.plotly_chart(fig_bar, use_container_width=True)

        # Evolução mensal por área (linhas)
        st.markdown("##### 📈 Evolução mensal por Área")
        df_evol = (df_ano_area.groupby(["mes_ano", "area_negocio"])["valor"]
                   .sum().reset_index())
        df_evol["mes_label"] = df_evol["mes_ano"].apply(
            lambda m: f"{nome_mes(m)[:3]}/{m[2:4]}"
        )
        fig_evol = px.line(
            df_evol.sort_values("mes_ano"),
            x="mes_label", y="valor", color="area_negocio",
            color_discrete_map={a: AREAS_NEGOCIO.get(a, {}).get("cor", "#616161") for a in por_area.index},
            markers=True,
            labels={"valor": "R$", "mes_label": "Mês", "area_negocio": "Área"},
        )
        fig_evol.update_layout(height=350, margin=dict(t=10, b=10))
        st.plotly_chart(fig_evol, use_container_width=True)

        # Drill-down: detalhamento por subcategoria
        st.markdown("##### 🔍 Detalhar uma área")
        area_escolhida = st.selectbox(
            "Escolha a área pra ver os fornecedores e subcategorias",
            por_area.index.tolist(),
            key="fin_visao_drill",
        )

        df_area = df_ano_area[df_ano_area["area_negocio"] == area_escolhida]
        df_area["subcategoria"] = df_area["subcategoria"].fillna("(sem subcategoria)")

        # Por subcategoria
        col_s, col_f = st.columns([1, 1])
        with col_s:
            st.markdown(f"**Por Subcategoria** ({area_escolhida})")
            por_sub = df_area.groupby("subcategoria")["valor"].sum().sort_values(ascending=False)
            rows_sub = []
            for sub, val in por_sub.items():
                rows_sub.append({
                    "Subcategoria": sub,
                    "Total": formatar_brl(val),
                    "%": f"{val / por_sub.sum() * 100:.1f}%",
                })
            st.dataframe(pd.DataFrame(rows_sub), use_container_width=True, hide_index=True)

        with col_f:
            st.markdown(f"**Por Fornecedor** ({area_escolhida})")
            por_forn_area = df_area.groupby("nome_fornecedor")["valor"].sum().sort_values(ascending=False)
            rows_fa = []
            for nome, val in por_forn_area.items():
                rows_fa.append({
                    "Fornecedor": nome,
                    "Total": formatar_brl(val),
                    "%": f"{val / por_forn_area.sum() * 100:.1f}%",
                })
            st.dataframe(pd.DataFrame(rows_fa), use_container_width=True, hide_index=True)

        # Alerta se tem fornecedor não classificado
        nao_class = df_ano_area[df_ano_area["area_negocio"] == "⚠️ NÃO CLASSIFICADO"]
        if not nao_class.empty:
            forn_nao_class = nao_class.groupby(["nome_fornecedor", "cnpj_fornecedor"])["valor"].sum()
            total_nc = nao_class["valor"].sum()
            with st.expander(
                f"⚠️ **{len(forn_nao_class)} fornecedor(es) sem classificação** ({formatar_brl(total_nc)}) — clique pra ver",
                expanded=False,
            ):
                st.caption("Vá em **🏢 Fornecedores** pra classificar esses fornecedores em uma área.")
                rows_nc = []
                for (nome, cnpj), val in forn_nao_class.items():
                    rows_nc.append({
                        "Fornecedor": nome,
                        "CNPJ": cnpj,
                        "Total": formatar_brl(val),
                    })
                st.dataframe(pd.DataFrame(rows_nc), use_container_width=True, hide_index=True)

    st.markdown("---")

    # ─── Por Fornecedor (Top 15) — agrupa por CNPJ pra evitar duplicações ─────
    st.markdown(f"### 🏢 Top Fornecedores — {ano_foco}")

    # Agrupa por CNPJ (chave única) e pega o nome MAIS CURTO (mais limpo) como display
    df_ano_forn = df_ano.copy()
    df_ano_forn["nome_len"] = df_ano_forn["nome_fornecedor"].fillna("").str.len()
    # Pra cada CNPJ, escolhe o nome mais curto (provavelmente o limpo)
    nomes_por_cnpj = (df_ano_forn.sort_values("nome_len")
                      .groupby("cnpj_fornecedor")["nome_fornecedor"]
                      .first().to_dict())

    # Pra cada CNPJ, pega também a área (snapshot do primeiro lançamento)
    areas_por_cnpj = (df_ano_area.groupby("cnpj_fornecedor")["area_negocio"]
                      .first().to_dict())

    por_forn = (df_ano.groupby("cnpj_fornecedor")["valor"]
                .sum().sort_values(ascending=False).head(15))

    rows = []
    for cnpj, val in por_forn.items():
        area = areas_por_cnpj.get(cnpj, "—")
        emoji = AREAS_NEGOCIO.get(area, {}).get("emoji", "❓") if area in AREAS_NEGOCIO else ""
        rows.append({
            "Fornecedor": nomes_por_cnpj.get(cnpj, "—"),
            "CNPJ": cnpj or "—",
            "Área": f"{emoji} {area}" if emoji else area,
            "Total": formatar_brl(val),
            "%": f"{val / total_ano * 100:.1f}%" if total_ano else "0%",
        })
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    st.markdown("---")

    # ─── Fornecedores mês a mês (tabela cruzada) ─────────────────────────
    st.markdown(f"### 🔁 Fornecedores mês a mês — {ano_foco}")
    st.caption(
        "Cada linha é um fornecedor. **Variação** = comparação do último mês com o anterior "
        "(considerando só meses com lançamento)."
    )

    if df_ano.empty:
        st.info("Sem dados pra esse ano.")
    else:
        meses_lista = sorted(df_ano["mes_ano"].unique())

        # Ordena por total desc, agrupando por CNPJ
        cnpjs_ordenados = (df_ano.groupby("cnpj_fornecedor")["valor"]
                           .sum().sort_values(ascending=False).index.tolist())

        linhas_tabela = []
        for cnpj in cnpjs_ordenados:
            lancs = df_ano[df_ano["cnpj_fornecedor"] == cnpj]
            por_mes = lancs.groupby("mes_ano")["valor"].sum().to_dict()
            row = {"Fornecedor": nomes_por_cnpj.get(cnpj, "—")}
            total = 0.0
            for m in meses_lista:
                v = float(por_mes.get(m, 0))
                row[f"{nome_mes(m)[:3]}/{m[2:4]}"] = formatar_brl(v) if v > 0 else "—"
                total += v
            row["Total"] = formatar_brl(total)

            # Variação: pega os 2 últimos meses COM lançamento desse fornecedor
            meses_com_valor = [(m, por_mes[m]) for m in meses_lista if por_mes.get(m, 0) > 0]
            if len(meses_com_valor) >= 2:
                ultimo = meses_com_valor[-1][1]
                penultimo = meses_com_valor[-2][1]
                if penultimo > 0:
                    pct = (ultimo - penultimo) / penultimo * 100
                    if abs(pct) < 0.5:
                        row["Variação"] = "0%"
                    elif pct > 0:
                        emoji = "🔺" if pct > 20 else "↗"
                        row["Variação"] = f"{emoji} +{pct:.1f}%"
                    else:
                        emoji = "🔻" if pct < -20 else "↘"
                        row["Variação"] = f"{emoji} {pct:.1f}%"
                else:
                    row["Variação"] = "—"
            else:
                row["Variação"] = "—"  # só 1 mês de dados, não dá pra comparar

            linhas_tabela.append(row)

        st.dataframe(pd.DataFrame(linhas_tabela), use_container_width=True, hide_index=True)

        # Total por mês
        totais = {"Fornecedor": "**TOTAL DO MÊS**"}
        soma_geral = 0.0
        valores_totais_mes = []
        for m in meses_lista:
            t = df_ano[df_ano["mes_ano"] == m]["valor"].sum()
            totais[f"{nome_mes(m)[:3]}/{m[2:4]}"] = formatar_brl(t)
            soma_geral += t
            if t > 0:
                valores_totais_mes.append(t)
        totais["Total"] = formatar_brl(soma_geral)

        # Variação do total geral
        if len(valores_totais_mes) >= 2:
            pct_geral = (valores_totais_mes[-1] - valores_totais_mes[-2]) / valores_totais_mes[-2] * 100
            if abs(pct_geral) < 0.5:
                totais["Variação"] = "0%"
            elif pct_geral > 0:
                emoji = "🔺" if pct_geral > 20 else "↗"
                totais["Variação"] = f"{emoji} +{pct_geral:.1f}%"
            else:
                emoji = "🔻" if pct_geral < -20 else "↘"
                totais["Variação"] = f"{emoji} {pct_geral:.1f}%"
        else:
            totais["Variação"] = "—"

        st.dataframe(pd.DataFrame([totais]), use_container_width=True, hide_index=True)

    st.markdown("---")

    # ─── Por Empresa LLE ─────────────────────────
    st.markdown(f"### 🏛️ Por Empresa LLE — {ano_foco}")

    por_emp = df_ano.groupby("empresa_lle", dropna=False)["valor"].sum().sort_values(ascending=False)
    if not por_emp.empty:
        col_e1, col_e2 = st.columns([2, 3])
        with col_e1:
            rows = []
            for emp, val in por_emp.items():
                rows.append({
                    "Empresa": emp or "—",
                    "Total": formatar_brl(val),
                    "%": f"{val / total_ano * 100:.1f}%" if total_ano else "0%",
                })
            st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)
        with col_e2:
            # Cores oficiais LLE
            cores_emp = {
                "PISA": "#041747",
                "KING": "#FAC318",
                "TRIO": "#0F8C3B",
                None: "#888888",
                "OUTRO": "#888888",
            }
            fig_emp = px.bar(
                x=[e or "—" for e in por_emp.index],
                y=por_emp.values,
                color=[e or "—" for e in por_emp.index],
                color_discrete_map={(e or "—"): cores_emp.get(e, "#888") for e in por_emp.index},
                text_auto=".2s",
                labels={"x": "Empresa", "y": "Total (R$)"},
            )
            fig_emp.update_layout(height=300, showlegend=False, margin=dict(t=10, b=10))
            fig_emp.update_traces(textposition="outside")
            st.plotly_chart(fig_emp, use_container_width=True)

    st.markdown("---")

    # ─── Projeção pro ano (orçamento próximo) ─────────────────────────
    st.markdown(f"### 🔮 Projeção e referência pra orçamento")

    meses_com_dado = df_ano[df_ano["valor"] > 0]["mes"].nunique()
    if meses_com_dado > 0:
        media_mensal = total_ano / meses_com_dado
        projecao_ano = media_mensal * 12

        col_p1, col_p2, col_p3 = st.columns(3)
        col_p1.metric("Meses com dados", meses_com_dado)
        col_p2.metric("Média mensal real", formatar_brl(media_mensal))
        col_p3.metric(
            f"Projeção fim de {ano_foco}",
            formatar_brl(projecao_ano),
            help="Média mensal × 12. Use como referência pra orçar o ano seguinte.",
        )

        st.info(
            f"💡 Pra orçar **{ano_foco + 1}**, considere a projeção de "
            f"**{formatar_brl(projecao_ano)}** como base. Inflação e novos "
            f"contratos devem ser adicionados separadamente."
        )

    # ─── Botão de export ─────────────────────────
    st.markdown("---")
    st.markdown("### 📥 Exportar relatório")

    if st.button("📊 Baixar Excel completo do ano", type="primary", key="btn_export_fin"):
        try:
            from io import BytesIO
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            # Aba 1: Lançamentos detalhados
            ws1 = wb.active
            ws1.title = "Lancamentos"
            headers = ["Mês", "Data", "Fornecedor", "CNPJ", "Categoria",
                       "Empresa", "NF nº", "Valor", "Descrição"]
            for c, h in enumerate(headers, 1):
                cell = ws1.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="041747", end_color="041747", fill_type="solid")

            for r, row in enumerate(df_ano.sort_values(["mes", "data_emissao"]).itertuples(), 2):
                ws1.cell(row=r, column=1, value=row.mes_ano)
                ws1.cell(row=r, column=2, value=str(row.data_emissao or ""))
                ws1.cell(row=r, column=3, value=row.nome_fornecedor)
                ws1.cell(row=r, column=4, value=row.cnpj_fornecedor)
                ws1.cell(row=r, column=5, value=row.categoria or "")
                ws1.cell(row=r, column=6, value=row.empresa_lle or "")
                ws1.cell(row=r, column=7, value=row.numero_nf or "")
                ws1.cell(row=r, column=8, value=float(row.valor))
                ws1.cell(row=r, column=9, value=row.descricao_servico or "")

            for col in ws1.columns:
                ws1.column_dimensions[col[0].column_letter].width = 22

            # Aba 2: Resumo mensal
            ws2 = wb.create_sheet("Resumo mensal")
            ws2.append(["Mês", "Total"])
            ws2["A1"].font = Font(bold=True)
            ws2["B1"].font = Font(bold=True)
            for m in range(1, 13):
                v = por_mes_foco[por_mes_foco["mes"] == m]["valor"].iloc[0] if not por_mes_foco.empty else 0
                ws2.append([nome_mes(f"2026-{m:02d}"), float(v)])

            # Aba 3: Por categoria
            ws3 = wb.create_sheet("Por categoria")
            ws3.append(["Categoria", "Total", "%"])
            for cat, val in por_cat.items():
                pct = (val / total_ano * 100) if total_ano else 0
                ws3.append([cat or "—", float(val), f"{pct:.1f}%"])

            # Aba 4: Top fornecedores
            ws4 = wb.create_sheet("Top fornecedores")
            ws4.append(["Fornecedor", "CNPJ", "Total", "%"])
            for (nome, cnpj), val in por_forn.items():
                pct = (val / total_ano * 100) if total_ano else 0
                ws4.append([nome, cnpj, float(val), f"{pct:.1f}%"])

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            st.download_button(
                label=f"⬇️ Baixar relatorio_financeiro_{ano_foco}.xlsx",
                data=buffer.getvalue(),
                file_name=f"relatorio_financeiro_{ano_foco}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        except Exception as e:
            st.error(f"Erro ao gerar Excel: {e}")
